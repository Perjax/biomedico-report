"""
módulo: excel_exporter.py
Propósito: Exportación del inventario biomédico a Excel con formato profesional.
Genera un archivo .xlsx con hoja de datos coloreada, hoja de resumen
y formato condicional automático.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Mapa de colores por estado (Excel HEX sin #)
# ---------------------------------------------------------------------------
COLOR_CABECERA = "1A237E"
COLOR_NORMAL = "C8E6C9"
COLOR_BAJO = "FFE0B2"
COLOR_CRITICO = "FFCDD2"
COLOR_AZUL_SUAVE = "E3F2FD"
COLOR_TEXTO_CABECERA = "FFFFFF"


def _aplicar_formato_cabecera(ws, fila: int, n_cols: int) -> None:
    """Aplica formato de cabecera institucional a la fila indicada."""
    for col in range(1, n_cols + 1):
        celda = ws.cell(row=fila, column=col)
        celda.font = Font(bold=True, color=COLOR_TEXTO_CABECERA, size=10)
        celda.fill = PatternFill("solid", fgColor=COLOR_CABECERA)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        borde = Side(style="thin", color="B0BEC5")
        celda.border = Border(left=borde, right=borde, top=borde, bottom=borde)


def _color_por_estado(estado: str, tipo: str) -> str:
    """Retorna el color HEX según el estado y tipo de campo."""
    if tipo == "bateria":
        mapa = {"CRÍTICO": COLOR_CRITICO, "BAJO": COLOR_BAJO, "NORMAL": COLOR_NORMAL}
    else:
        mapa = {"VENCIDA": COLOR_CRITICO, "PRÓXIMA": COLOR_BAJO, "VIGENTE": COLOR_NORMAL}
    return mapa.get(estado, "FFFFFF")


def _ajustar_anchos(ws) -> None:
    """Ajusta automáticamente el ancho de columnas según el contenido."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for celda in col:
            try:
                if celda.value:
                    max_len = max(max_len, len(str(celda.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)


def _hoja_inventario(writer, df: pd.DataFrame) -> None:
    """
    Escribe la hoja principal de inventario con formato profesional.

    Args:
        writer: ExcelWriter activo.
        df: DataFrame con datos de equipos.
    """
    # Preparar DataFrame de salida (sin columna booleana interna)
    df_export = df.copy()
    df_export["Alerta Crítica"] = df_export["Alerta Crítica"].apply(
        lambda x: "⚠ CRÍTICO" if x else "✓ OK"
    )

    df_export.to_excel(writer, sheet_name="Inventario", index=False, startrow=2)
    ws = writer.sheets["Inventario"]

    # Título superior
    ws.merge_cells("A1:H1")
    celda_titulo = ws["A1"]
    celda_titulo.value = "SISTEMA DE GESTIÓN DE EQUIPOS BIOMÉDICOS — Inventario"
    celda_titulo.font = Font(bold=True, size=13, color=COLOR_TEXTO_CABECERA)
    celda_titulo.fill = PatternFill("solid", fgColor=COLOR_CABECERA)
    celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Formato de cabecera de columnas (fila 3 = startrow+1)
    _aplicar_formato_cabecera(ws, fila=3, n_cols=len(df_export.columns))
    ws.row_dimensions[3].height = 22

    # Formato de filas de datos
    idx_bat = df_export.columns.tolist().index("Estado Batería") + 1
    idx_cal = df_export.columns.tolist().index("Estado Calibración") + 1
    idx_alerta = df_export.columns.tolist().index("Alerta Crítica") + 1

    for i, (_, row) in enumerate(df_export.iterrows(), start=4):
        color_fila = "FFFFFF" if i % 2 == 0 else "F5F7FA"
        borde = Side(style="thin", color="CFD8DC")

        for j, val in enumerate(row, start=1):
            celda = ws.cell(row=i, column=j)
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.fill = PatternFill("solid", fgColor=color_fila)
            celda.border = Border(left=borde, right=borde, top=borde, bottom=borde)

        # Colorear columnas de estado
        ws.cell(row=i, column=idx_bat).fill = PatternFill(
            "solid", fgColor=_color_por_estado(row["Estado Batería"], "bateria")
        )
        ws.cell(row=i, column=idx_cal).fill = PatternFill(
            "solid", fgColor=_color_por_estado(row["Estado Calibración"], "calibracion")
        )
        if "CRÍTICO" in str(row["Alerta Crítica"]):
            ws.cell(row=i, column=idx_alerta).fill = PatternFill("solid", fgColor=COLOR_CRITICO)
            ws.cell(row=i, column=idx_alerta).font = Font(bold=True, color="C62828")

        ws.row_dimensions[i].height = 18

    _ajustar_anchos(ws)
    ws.freeze_panes = "A4"  # Fijar cabecera al hacer scroll


def _hoja_resumen(writer, df: pd.DataFrame) -> None:
    """
    Escribe la hoja de resumen con métricas y estadísticas clave.

    Args:
        writer: ExcelWriter activo.
        df: DataFrame con datos de equipos.
    """
    resumen = {
        "Total de Equipos": len(df),
        "Equipos Críticos": int(df["Alerta Crítica"].sum()),
        "Batería CRÍTICA (≤30%)": int((df["Estado Batería"] == "CRÍTICO").sum()),
        "Batería BAJA (31-60%)": int((df["Estado Batería"] == "BAJO").sum()),
        "Batería NORMAL (>60%)": int((df["Estado Batería"] == "NORMAL").sum()),
        "Calibración VENCIDA (>90d)": int((df["Estado Calibración"] == "VENCIDA").sum()),
        "Calibración PRÓXIMA (61-90d)": int((df["Estado Calibración"] == "PRÓXIMA").sum()),
        "Calibración VIGENTE (≤60d)": int((df["Estado Calibración"] == "VIGENTE").sum()),
        "Batería Promedio (%)": round(df["Batería (%)"].mean(), 1),
        "Días Promedio sin Calibrar": round(df["Días desde Calibración"].mean(), 1),
    }

    df_resumen = pd.DataFrame(list(resumen.items()), columns=["Métrica", "Valor"])
    df_resumen.to_excel(writer, sheet_name="Resumen", index=False, startrow=2)
    ws = writer.sheets["Resumen"]

    ws.merge_cells("A1:B1")
    ws["A1"].value = "RESUMEN EJECUTIVO — Gestión de Equipos Biomédicos"
    ws["A1"].font = Font(bold=True, size=12, color=COLOR_TEXTO_CABECERA)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_CABECERA)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    _aplicar_formato_cabecera(ws, fila=3, n_cols=2)

    for i in range(4, 4 + len(df_resumen)):
        color = "FFFFFF" if i % 2 == 0 else "F5F7FA"
        for j in [1, 2]:
            celda = ws.cell(row=i, column=j)
            borde = Side(style="thin", color="CFD8DC")
            celda.border = Border(left=borde, right=borde, top=borde, bottom=borde)
            celda.fill = PatternFill("solid", fgColor=color)
            celda.alignment = Alignment(horizontal="center" if j == 2 else "left",
                                        vertical="center")

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14


def exportar_excel(df: pd.DataFrame, ruta_salida: str = "reports/reporte_biomedico.xlsx") -> str:
    """
    Exporta el inventario biomédico a un archivo Excel con dos hojas:
    'Inventario' (tabla detallada) y 'Resumen' (métricas ejecutivas).

    Args:
        df: DataFrame con datos de equipos.
        ruta_salida: Ruta del archivo Excel de salida.

    Returns:
        Ruta del archivo Excel generado.
    """
    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        _hoja_inventario(writer, df)
        _hoja_resumen(writer, df)

    return ruta_salida
